"""Source-specific adapters for the Aptafind Silver harmonization layer."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable

import openpyxl
import pandas as pd

from aptafind.data.aptabench_profile import normalize_origin
from aptafind.data.harmonization_schema import SourceMeasurement


@dataclass
class AdapterResult:
    """Records and source-local audit values returned by one adapter."""

    records: list[SourceMeasurement]
    audit: dict[str, Any]


def _text(value: Any) -> str | None:
    if value is None or pd.isna(value):
        return None
    text = " ".join(str(value).strip().split())
    return text or None


def _float(value: Any) -> float | None:
    if value is None or pd.isna(value):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _integer(value: Any) -> int | None:
    number = _float(value)
    return int(number) if number is not None else None


def _publication_identifier(value: Any) -> str | None:
    text = _text(value)
    return normalize_origin(text) if text is not None else None


def adapt_aptabench(path: Path, dataset_id: str) -> AdapterResult:
    """Adapt the frozen AptaBench CSV, retaining ssDNA binary supervision."""

    frame = pd.read_csv(path)
    expected = [
        "type",
        "sequence",
        "canonical_smiles",
        "pKd_value",
        "label",
        "buffer",
        "origin",
        "source",
    ]
    if list(frame.columns) != expected:
        raise ValueError(
            f"Unexpected AptaBench columns in {path}: {list(frame.columns)}"
        )
    if not set(frame["label"].dropna().astype(int)).issubset({0, 1}):
        raise ValueError("AptaBench labels must be binary values 0 or 1.")

    dna = frame.loc[frame["type"].eq("DNA")]
    records: list[SourceMeasurement] = []
    for source_index, row in dna.iterrows():
        label = int(row["label"])
        pkd = _float(row["pKd_value"])
        source_lineage = _text(row["source"])
        if source_lineage == "Specificity":
            evidence_type = "measured_cross_reactivity"
        elif pkd is not None:
            evidence_type = "measured_affinity"
        else:
            evidence_type = "published_binary_label"
        records.append(
            SourceMeasurement(
                source_dataset=dataset_id,
                source_file=path.name,
                source_sheet=None,
                source_row=int(source_index) + 2,
                source_record_id=f"csv-row:{int(source_index) + 2}",
                source_aptamer_id=None,
                sequence_raw=_text(row["sequence"]),
                polymer_type="DNA",
                sequence_role="published_sequence",
                source_target_id=None,
                target_name_raw=None,
                target_smiles_raw=_text(row["canonical_smiles"]),
                publication_id=_publication_identifier(row["origin"]),
                measurement_type="pkd" if pkd is not None else "binary_binding",
                measurement_value=pkd if pkd is not None else float(label),
                measurement_unit="pKd" if pkd is not None else "binary",
                binding_label="positive" if label == 1 else "negative",
                evidence_type=evidence_type,
                is_target_measurement=bool(label),
                assay=source_lineage,
                buffer=_text(row["buffer"]),
                details={"aptabench_source": source_lineage},
            )
        )
    return AdapterResult(
        records=records,
        audit={
            "source_rows": int(len(frame)),
            "dna_rows_emitted": int(len(records)),
            "positive_rows": int((dna["label"] == 1).sum()),
            "negative_rows": int((dna["label"] == 0).sum()),
        },
    )


def adapt_utexas(path: Path, dataset_id: str) -> AdapterResult:
    """Adapt the UTexas workbook's unmodified ssDNA records."""

    frame = pd.read_excel(path, sheet_name="Dataset")
    required = {
        "Type of Nucleic Acid",
        "Name of Aptamer",
        "Target ",
        "Aptamer Sequence",
        "Kd (nM)",
        "Affinity",
        "Journal DOI",
        "Link to PubMed Entry",
        "Binding Buffer/Conditions",
        "Serial Number",
    }
    missing = required - set(frame.columns)
    if missing:
        raise ValueError(f"UTexas workbook is missing columns: {sorted(missing)}")

    dna = frame.loc[frame["Type of Nucleic Acid"].eq("ssDNA")]
    records: list[SourceMeasurement] = []
    for source_index, row in dna.iterrows():
        kd_nm = _float(row["Kd (nM)"])
        doi = _text(row["Journal DOI"])
        publication = _publication_identifier(
            doi if doi is not None else row["Link to PubMed Entry"]
        )
        source_aptamer_id = _text(row["Name of Aptamer"])
        serial = _text(row["Serial Number"])
        if serial is not None:
            source_aptamer_id = f"{serial}:{source_aptamer_id or 'unnamed'}"
        records.append(
            SourceMeasurement(
                source_dataset=dataset_id,
                source_file=path.name,
                source_sheet="Dataset",
                source_row=int(source_index) + 2,
                source_record_id=f"Dataset-row:{int(source_index) + 2}",
                source_aptamer_id=source_aptamer_id,
                sequence_raw=_text(row["Aptamer Sequence"]),
                polymer_type="DNA",
                sequence_role="published_sequence",
                source_target_id=None,
                target_name_raw=_text(row["Target "]),
                target_smiles_raw=None,
                publication_id=publication,
                measurement_type="kd" if kd_nm is not None else "published_candidate",
                measurement_value=kd_nm,
                measurement_unit="nM" if kd_nm is not None else None,
                binding_label="positive",
                evidence_type=(
                    "measured_affinity" if kd_nm is not None else "published_candidate"
                ),
                is_target_measurement=True,
                assay="UTexas curated published assay",
                buffer=_text(row["Binding Buffer/Conditions"]),
                value_qualifier=_text(row["Affinity"]),
                details={"source_polymer_label": "ssDNA"},
            )
        )
    return AdapterResult(
        records=records,
        audit={
            "source_rows": int(len(frame)),
            "ssdna_rows_emitted": int(len(records)),
            "rows_with_numeric_kd_nm": int(dna["Kd (nM)"].notna().sum()),
        },
    )


def adapt_aptadb(
    interaction_path: Path,
    aptamer_path: Path,
    molecule_path: Path,
    dataset_id: str,
) -> AdapterResult:
    """Adapt AptaDB's DNA-small-molecule interaction subset."""

    interactions = pd.read_csv(interaction_path, encoding="latin_1")
    aptamers = pd.read_csv(aptamer_path, encoding="latin_1")
    molecules = pd.read_csv(molecule_path, encoding="utf_8")

    joined = interactions.merge(
        aptamers[
            ["Apta_index", "Aptamer Chemistry", "Aptamer description", "Assay"]
        ],
        on="Apta_index",
        how="left",
        validate="many_to_one",
    )
    molecules = molecules.copy()
    molecules["Pubchem ID"] = molecules["Pubchem ID"].astype(str)
    joined["TargetID"] = joined["TargetID"].astype(str)
    joined = joined.merge(
        molecules[["Pubchem ID", "Titles", "InChIKeys", "IUPAC Names"]],
        left_on="TargetID",
        right_on="Pubchem ID",
        how="left",
        validate="many_to_one",
    )
    selected = joined.loc[
        joined["Target chemistry"].eq("Molecule")
        & joined["Aptamer Chemistry"].eq("DNA")
    ]
    if bool(selected["Titles"].isna().any()):
        raise ValueError("AptaDB molecule interactions did not all join to molecule.csv.")

    records: list[SourceMeasurement] = []
    for source_index, row in selected.iterrows():
        source_row = _integer(row["Index"]) or int(source_index) + 2
        records.append(
            SourceMeasurement(
                source_dataset=dataset_id,
                source_file=interaction_path.name,
                source_sheet=None,
                source_row=source_row,
                source_record_id=f"interaction:{source_row}",
                source_aptamer_id=_text(row["Apta_index"]),
                sequence_raw=_text(row["Sequence"]),
                polymer_type="DNA",
                sequence_role="published_sequence",
                source_target_id=_text(row["TargetID"]),
                target_name_raw=_text(row["Titles"]),
                target_smiles_raw=None,
                publication_id=_publication_identifier(row["Reference(pubmed ID)"]),
                measurement_type="published_interaction",
                measurement_value=None,
                measurement_unit=None,
                binding_label="positive",
                evidence_type="published_candidate",
                is_target_measurement=True,
                assay=_text(row["Assay"]),
                buffer=_text(row["Binding Conditions/Buffer"]),
                details={
                    "aptamer_description": _text(row["Aptamer description"]),
                    "target_inchikey": _text(row["InChIKeys"]),
                    "target_iupac_name": _text(row["IUPAC Names"]),
                },
            )
        )
    return AdapterResult(
        records=records,
        audit={
            "interaction_rows": int(len(interactions)),
            "dna_small_molecule_rows_emitted": int(len(records)),
            "unique_pubchem_targets": int(selected["TargetID"].nunique()),
        },
    )


N2A2_TARGETS = {
    "Kyn": "kynurenine",
    "KA": "kynurenic acid",
    "3HK": "3-hydroxykynurenine",
    "3HA": "3-hydroxyanthranilic acid",
    "XA": "xanthurenic acid",
}


def adapt_n2a2(path: Path, dataset_id: str) -> AdapterResult:
    """Melt the N2A2 sequence-level z-score summary into target measurements."""

    frame = pd.read_excel(path, sheet_name="mean_z_score_df")
    required = {"seq", "replicate"}
    for code in N2A2_TARGETS:
        required.update({f"{code}_mean", f"{code}_stdev"})
    missing = required - set(frame.columns)
    if missing:
        raise ValueError(f"N2A2 workbook is missing columns: {sorted(missing)}")

    records: list[SourceMeasurement] = []
    positive_count = 0
    for source_index, row in frame.iterrows():
        replicate_count = _integer(row["replicate"])
        for code, target_name in N2A2_TARGETS.items():
            value = _float(row[f"{code}_mean"])
            error = _float(row[f"{code}_stdev"])
            is_screen_positive = value is not None and value >= 2.576
            positive_count += int(is_screen_positive)
            records.append(
                SourceMeasurement(
                    source_dataset=dataset_id,
                    source_file=path.name,
                    source_sheet="mean_z_score_df",
                    source_row=int(source_index) + 2,
                    source_record_id=(
                        f"mean_z_score_df:{int(source_index) + 2}:{code}"
                    ),
                    source_aptamer_id=None,
                    sequence_raw=_text(row["seq"]),
                    polymer_type="DNA",
                    sequence_role="variable_region",
                    source_target_id=code,
                    target_name_raw=target_name,
                    target_smiles_raw=None,
                    publication_id="doi:10.1073/pnas.2119945119",
                    measurement_type="screen_z_score",
                    measurement_value=value,
                    measurement_unit="z_score",
                    measurement_error=error,
                    measurement_error_unit="z_score_stdev",
                    binding_label=(
                        "screen_positive"
                        if is_screen_positive
                        else "screen_below_threshold"
                    ),
                    evidence_type="measured_cross_reactivity",
                    is_target_measurement=None,
                    assay="N2A2 modified Illumina MiSeq specificity screen",
                    replicate_count=replicate_count,
                    details={"screen_threshold_z": 2.576},
                    transform_notes=(
                        "Threshold flag records z-score evidence only; it does not "
                        "apply the publication's separate specificity-ratio rule."
                    ),
                )
            )
    return AdapterResult(
        records=records,
        audit={
            "sequence_summary_rows": int(len(frame)),
            "target_measurements_emitted": int(len(records)),
            "z_score_threshold_flags": int(positive_count),
            "unique_sequences": int(frame["seq"].nunique()),
        },
    )


def _target_marker(value: Any) -> bool:
    text = _text(value)
    return text is not None and "(target)" in text.casefold()


def _strip_target_marker(value: Any) -> str | None:
    text = _text(value)
    if text is None:
        return None
    return text.replace("(Target)", "").replace("(target)", "").strip()


def adapt_xiao_specificity(path: Path, dataset_id: str) -> AdapterResult:
    """Convert Xiao-lab cross-reactivity panels from wide blocks to long rows."""

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    records: list[SourceMeasurement] = []
    target_measurements = 0
    non_target_measurements = 0
    aptamer_ids: set[str] = set()

    for sheet in workbook.worksheets:
        if sheet.title.casefold() == "readme":
            continue
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        candidate_header_rows = range(min(5, len(rows)))
        header_row_index = max(
            candidate_header_rows,
            key=lambda index: sum(_target_marker(value) for value in rows[index]),
        )
        target_starts = [
            index
            for index, value in enumerate(rows[header_row_index])
            if _target_marker(value)
        ]
        if not target_starts or header_row_index == 0:
            raise ValueError(f"Could not locate specificity blocks in {sheet.title}.")
        aptamer_header = rows[header_row_index - 1]

        for block_index, start in enumerate(target_starts):
            end = (
                target_starts[block_index + 1]
                if block_index + 1 < len(target_starts)
                else len(rows[header_row_index])
            )
            intended_target = _strip_target_marker(rows[header_row_index][start])
            aptamer_columns = [
                column
                for column in range(start + 1, min(end, len(aptamer_header)))
                if _text(aptamer_header[column]) is not None
            ]
            if not aptamer_columns:
                raise ValueError(
                    f"Specificity block {sheet.title}:{intended_target} has no aptamers."
                )

            for row_index in range(header_row_index, len(rows)):
                row = rows[row_index]
                ligand = _text(row[start]) if start < len(row) else None
                if ligand is None or ligand.casefold().startswith("specificity score"):
                    continue
                ligand = _strip_target_marker(ligand)
                is_target = row_index == header_row_index
                for column in aptamer_columns:
                    value = _float(row[column]) if column < len(row) else None
                    if value is None:
                        continue
                    aptamer_id = _text(aptamer_header[column])
                    if aptamer_id is None:
                        continue
                    aptamer_ids.add(aptamer_id)
                    if is_target:
                        binding_label = "positive"
                        target_measurements += 1
                    elif value == 0:
                        binding_label = "measured_nonresponse"
                        non_target_measurements += 1
                    else:
                        binding_label = "cross_reactive"
                        non_target_measurements += 1
                    records.append(
                        SourceMeasurement(
                            source_dataset=dataset_id,
                            source_file=path.name,
                            source_sheet=sheet.title,
                            source_row=row_index + 1,
                            source_record_id=(
                                f"{sheet.title}:{intended_target}:{aptamer_id}:"
                                f"{row_index + 1}:{ligand}"
                            ),
                            source_aptamer_id=aptamer_id,
                            sequence_raw=None,
                            polymer_type="DNA",
                            sequence_role="published_identifier_only",
                            source_target_id=None,
                            target_name_raw=ligand,
                            target_smiles_raw=None,
                            publication_id="doi:10.1093/nar/gkaf219",
                            measurement_type="cross_reactivity",
                            measurement_value=value,
                            measurement_unit="fraction_of_target_response",
                            binding_label=binding_label,
                            evidence_type="measured_cross_reactivity",
                            is_target_measurement=is_target,
                            assay="exonuclease digestion cross-reactivity assay",
                            details={"intended_target": intended_target},
                            transform_notes=(
                                "The source readme states that negative "
                                "cross-reactivity values were replaced with zero."
                            ),
                        )
                    )

    return AdapterResult(
        records=records,
        audit={
            "aptamer_identifiers": int(len(aptamer_ids)),
            "target_measurements": int(target_measurements),
            "non_target_measurements": int(non_target_measurements),
            "total_measurements": int(len(records)),
        },
    )


def adapt_xiao_itc(path: Path, dataset_id: str) -> AdapterResult:
    """Convert the Xiao-lab block-formatted ITC master sheet to pair records."""

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["All ITC Data"]
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 4:
        raise ValueError("Xiao ITC sheet is unexpectedly short.")
    group_header = rows[0]
    group_starts = [
        index for index, value in enumerate(group_header) if _text(value) is not None
    ]

    records: list[SourceMeasurement] = []
    aptamer_ids: list[str] = []
    for group_index, start in enumerate(group_starts):
        end = (
            group_starts[group_index + 1]
            if group_index + 1 < len(group_starts)
            else len(group_header)
        )
        group_name = _text(group_header[start])
        if group_name is None:
            continue
        misc = group_name.casefold() == "misc"
        minimum_width = 8 if misc else 7
        if end - start < minimum_width:
            raise ValueError(f"ITC block {group_name!r} is narrower than expected.")

        for row_index in range(3, len(rows)):
            row = rows[row_index]
            aptamer_id = _text(row[start]) if start < len(row) else None
            if aptamer_id is None:
                continue
            if misc:
                target_name = _text(row[start + 1])
                kd_column = start + 2
            else:
                target_name = group_name
                kd_column = start + 1
            kd_nm = _float(row[kd_column]) if kd_column < len(row) else None
            if target_name is None or kd_nm is None:
                continue
            delta_h_column = kd_column + 1
            delta_s_column = kd_column + 2
            t_delta_s_column = kd_column + 3
            binding_sites_column = kd_column + 4
            reference_column = kd_column + 5
            aptamer_ids.append(aptamer_id)
            records.append(
                SourceMeasurement(
                    source_dataset=dataset_id,
                    source_file=path.name,
                    source_sheet="All ITC Data",
                    source_row=row_index + 1,
                    source_record_id=(
                        f"All ITC Data:{group_name}:{aptamer_id}:{row_index + 1}"
                    ),
                    source_aptamer_id=aptamer_id,
                    sequence_raw=None,
                    polymer_type="DNA",
                    sequence_role="published_identifier_only",
                    source_target_id=None,
                    target_name_raw=target_name,
                    target_smiles_raw=None,
                    publication_id="doi:10.1093/nar/gkaf219",
                    measurement_type="itc_thermodynamics",
                    measurement_value=kd_nm,
                    measurement_unit="nM",
                    binding_label="positive",
                    evidence_type="measured_affinity",
                    is_target_measurement=True,
                    assay="isothermal titration calorimetry",
                    details={
                        "delta_h_bind_kcal_per_mol": _float(row[delta_h_column]),
                        "delta_s_bind_cal_per_mol_k": _float(row[delta_s_column]),
                        "t_delta_s_bind_kcal_per_mol": _float(row[t_delta_s_column]),
                        "binding_sites": _float(row[binding_sites_column]),
                        "source_reference": _text(row[reference_column]),
                        "source_target_group": group_name,
                    },
                )
            )

    return AdapterResult(
        records=records,
        audit={
            "itc_pair_records": int(len(records)),
            "target_groups": int(len(group_starts)),
            "distinct_aptamer_id_strings": int(len(set(aptamer_ids))),
        },
    )


ADAPTERS: dict[str, Callable[..., AdapterResult]] = {
    "aptabench": adapt_aptabench,
    "utexas": adapt_utexas,
    "aptadb": adapt_aptadb,
    "n2a2": adapt_n2a2,
    "xiao_specificity": adapt_xiao_specificity,
    "xiao_itc": adapt_xiao_itc,
}
